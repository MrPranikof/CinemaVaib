from core.database import query
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile


class ReportsModel:

    @staticmethod
    def get_daily_sales_report(days=30):
        """Отчет по ежедневным продажам (только актуальные билеты)"""
        sql = f"""
            SELECT 
                DATE(t.purchase_date) as date,
                COUNT(*) as tickets_sold,
                SUM(t.final_price) as revenue,
                AVG(t.final_price) as avg_ticket_price
            FROM ticket t
            WHERE t.purchase_date >= CURRENT_DATE - INTERVAL '{days} days'
            AND t.session_id IN (
                SELECT session_id FROM session 
                WHERE session_time >= CURRENT_DATE - INTERVAL '{days} days'
            )
            GROUP BY DATE(t.purchase_date)
            ORDER BY date DESC
        """
        return query(sql) or []

    @staticmethod
    def get_movies_popularity_report():
        """Отчет по популярности фильмов (только актуальные данные)"""
        sql = """
            SELECT 
                m.movie_id,
                m.title,
                COUNT(t.ticket_id) as tickets_sold,
                SUM(t.final_price) as revenue,
                AVG(t.final_price) as avg_ticket_price,
                COUNT(DISTINCT s.session_id) as sessions_count,
                COALESCE(AVG(r.rating), 0) as avg_rating
            FROM movies m
            LEFT JOIN session s ON m.movie_id = s.movie_id 
                AND s.session_time >= CURRENT_DATE - INTERVAL '365 days'  -- Только актуальные сеансы
            LEFT JOIN ticket t ON s.session_id = t.session_id
            LEFT JOIN review r ON m.movie_id = r.movie_id
            WHERE m.movie_id IN (SELECT DISTINCT movie_id FROM session)  -- Только фильмы с сеансами
            GROUP BY m.movie_id, m.title
            ORDER BY tickets_sold DESC, revenue DESC
        """
        return query(sql) or []

    @staticmethod
    def get_halls_utilization_report(days=30):
        """Отчет по загрузке залов (только актуальные сеансы)"""
        sql = f"""
            SELECT 
                h.hall_id,
                h.hall_name,
                h.hall_number,
                COUNT(s.session_id) as total_sessions,
                COUNT(t.ticket_id) as tickets_sold,
                SUM(t.final_price) as revenue,
                COUNT(DISTINCT s.movie_id) as unique_movies,
                ROUND(
                    COUNT(t.ticket_id) * 100.0 / 
                    NULLIF(
                        (SELECT COUNT(*) FROM seat WHERE hall_id = h.hall_id) * 
                        COUNT(DISTINCT s.session_id), 
                    0), 
                2) as avg_occupancy_percent
            FROM hall h
            LEFT JOIN session s ON h.hall_id = s.hall_id 
                AND s.session_time >= CURRENT_DATE - INTERVAL '{days} days'
                AND s.session_time <= CURRENT_DATE + INTERVAL '1 day'  -- Только прошедшие и ближайшие сеансы
            LEFT JOIN ticket t ON s.session_id = t.session_id
            GROUP BY h.hall_id, h.hall_name, h.hall_number
            ORDER BY revenue DESC NULLS LAST
        """
        return query(sql) or []

    @staticmethod
    def get_users_activity_report(days=30):
        """Отчет по активности пользователей (только актуальные покупки)"""
        sql = f"""
            SELECT 
                u.user_id,
                u.login,
                u.email,
                r.role_name,
                u.created_at,
                COUNT(t.ticket_id) as tickets_bought,
                SUM(t.final_price) as total_spent,
                COUNT(rev.review_id) as reviews_written,
                MAX(t.purchase_date) as last_activity
            FROM users u
            LEFT JOIN roles r ON u.role_id = r.role_id
            LEFT JOIN ticket t ON u.user_id = t.user_id 
                AND t.purchase_date >= CURRENT_DATE - INTERVAL '{days} days'
                AND t.session_id IN (
                    SELECT session_id FROM session 
                    WHERE session_time >= CURRENT_DATE - INTERVAL '{days} days'
                )
            LEFT JOIN review rev ON u.user_id = rev.user_id
            WHERE u.status = 'Active'  -- Только активные пользователи
            GROUP BY u.user_id, u.login, u.email, r.role_name, u.created_at
            ORDER BY total_spent DESC NULLS LAST, tickets_bought DESC
        """
        return query(sql) or []

    @staticmethod
    def get_financial_summary_report(days=30):
        """Финансовый отчет (только актуальные данные)"""
        try:
            # Числовые показатели (только актуальные билеты)
            numeric_sql = f"""
                SELECT 
                    COALESCE(SUM(t.final_price), 0) as total_revenue,
                    COALESCE(COUNT(*), 0) as total_tickets,
                    COALESCE(AVG(t.final_price), 0) as avg_ticket_price,
                    COALESCE(COUNT(DISTINCT user_id), 0) as unique_customers
                FROM ticket t
                WHERE t.purchase_date >= CURRENT_DATE - INTERVAL '{days} days'
                AND t.session_id IN (
                    SELECT session_id FROM session 
                    WHERE session_time >= CURRENT_DATE - INTERVAL '{days} days'
                )
            """
            numeric_result = query(numeric_sql)

            # Самый популярный фильм (только актуальные сеансы)
            popular_movie_sql = f"""
                SELECT m.title 
                FROM movies m
                JOIN session s ON m.movie_id = s.movie_id
                JOIN ticket t ON s.session_id = t.session_id
                WHERE t.purchase_date >= CURRENT_DATE - INTERVAL '{days} days'
                AND s.session_time >= CURRENT_DATE - INTERVAL '{days} days'
                GROUP BY m.movie_id, m.title
                ORDER BY COUNT(*) DESC
                LIMIT 1
            """
            popular_movie_result = query(popular_movie_sql)

            # Формируем результат
            if numeric_result:
                total_revenue, total_tickets, avg_ticket_price, unique_customers = numeric_result[0]
                popular_movie = popular_movie_result[0][0] if popular_movie_result else "Нет данных"

                return [
                    ("Общая выручка", total_revenue),
                    ("Количество билетов", total_tickets),
                    ("Средний чек", avg_ticket_price),
                    ("Уникальных клиентов", unique_customers),
                    ("Самый популярный фильм", popular_movie)
                ]
            else:
                return [
                    ("Общая выручка", 0),
                    ("Количество билетов", 0),
                    ("Средний чек", 0),
                    ("Уникальных клиентов", 0),
                    ("Самый популярный фильм", "Нет данных")
                ]

        except Exception as e:
            print(f"Ошибка финансового отчета: {e}")
            return []

    @staticmethod
    def create_excel_report(report_data, headers, title, filename_suffix=""):
        """Создать Excel файл с отчетом"""
        try:
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет"

            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')

            # Заголовок отчета
            ws.merge_cells('A1:H1')
            ws['A1'] = f"Отчет: {title}"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align

            # Дата генерации
            ws.merge_cells('A2:H2')
            ws['A2'] = f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws['A2'].alignment = center_align

            # Заголовки столбцов
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            # Данные
            for row, data_row in enumerate(report_data, 5):
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if isinstance(value, (int, float)) and col > 1:  # Числовые данные выравниваем по правому краю
                        cell.alignment = Alignment(horizontal='right')

            # Автоматическая ширина столбцов
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Создаем временный файл
            temp_dir = tempfile.gettempdir()
            filename = f"report_{filename_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(temp_dir, filename)

            wb.save(filepath)
            return filepath

        except Exception as e:
            print(f"Ошибка создания Excel отчета: {e}")
            return None

    @staticmethod
    def export_daily_sales_report(days=30):
        """Экспорт отчета по продажам в Excel"""
        data = ReportsModel.get_daily_sales_report(days)
        headers = ["Дата", "Билетов продано", "Выручка (руб.)", "Средний чек (руб.)"]
        title = f"Ежедневные продажи за {days} дней"

        # Форматируем данные для Excel
        formatted_data = []
        for row in data:
            formatted_data.append([
                row[0].strftime('%d.%m.%Y') if row[0] else '',
                row[1] or 0,
                float(row[2] or 0),
                float(row[3] or 0)
            ])

        return ReportsModel.create_excel_report(formatted_data, headers, title, "sales")

    @staticmethod
    def export_movies_popularity_report():
        """Экспорт отчета по популярности фильмов в Excel"""
        data = ReportsModel.get_movies_popularity_report()
        headers = ["ID", "Название фильма", "Билетов продано", "Выручка (руб.)",
                   "Средний чек (руб.)", "Кол-во сеансов", "Средний рейтинг"]

        formatted_data = []
        for row in data:
            formatted_data.append([
                row[0],
                row[1],
                row[2] or 0,
                float(row[3] or 0),
                float(row[4] or 0),
                row[5] or 0,
                float(row[6] or 0)
            ])

        return ReportsModel.create_excel_report(formatted_data, headers, "Популярность фильмов", "movies")

    @staticmethod
    def export_halls_utilization_report(days=30):
        """Экспорт отчета по загрузке залов в Excel"""
        data = ReportsModel.get_halls_utilization_report(days)
        headers = ["ID", "Название зала", "Номер", "Всего сеансов", "Билетов продано",
                   "Выручка (руб.)", "Уникальных фильмов", "Загрузка (%)"]

        formatted_data = []
        for row in data:
            formatted_data.append([
                row[0],
                row[1],
                row[2],
                row[3] or 0,
                row[4] or 0,
                float(row[5] or 0),
                row[6] or 0,
                float(row[7] or 0)
            ])

        return ReportsModel.create_excel_report(formatted_data, headers, f"Загрузка залов за {days} дней", "halls")

    @staticmethod
    def export_users_activity_report(days=30):
        """Экспорт отчета по активности пользователей в Excel"""
        data = ReportsModel.get_users_activity_report(days)
        headers = ["ID", "Логин", "Email", "Роль", "Дата регистрации",
                   "Куплено билетов", "Потрачено (руб.)", "Написано отзывов", "Последняя активность"]

        formatted_data = []
        for row in data:
            formatted_data.append([
                row[0],
                row[1],
                row[2],
                row[3],
                row[4].strftime('%d.%m.%Y') if row[4] else '',
                row[5] or 0,
                float(row[6] or 0),
                row[7] or 0,
                row[8].strftime('%d.%m.%Y %H:%M') if row[8] else 'Нет активности'
            ])

        return ReportsModel.create_excel_report(formatted_data, headers, f"Активность пользователей за {days} дней",
                                                "users")

    @staticmethod
    def export_financial_summary_report(days=30):
        """Экспорт финансового отчета в Excel"""
        data = ReportsModel.get_financial_summary_report(days)
        headers = ["Показатель", "Значение"]

        formatted_data = []
        for row in data:
            if row[0] == 'Общая выручка' and row[1]:
                value = f"{float(row[1]):,.2f} руб."
            elif row[0] == 'Средний чек' and row[1]:
                value = f"{float(row[1]):,.2f} руб."
            else:
                value = row[1] or "Нет данных"
            formatted_data.append([row[0], value])

        return ReportsModel.create_excel_report(formatted_data, headers, f"Финансовый отчет за {days} дней",
                                                "financial")

    @staticmethod
    def get_realtime_stats():
        """Статистика в реальном времени (только актуальные данные)"""
        sql = """
            SELECT 
                (SELECT COUNT(*) FROM movies WHERE movie_id IN 
                    (SELECT DISTINCT movie_id FROM session WHERE session_time >= CURRENT_DATE)) as active_movies,
                (SELECT COUNT(*) FROM session WHERE session_time >= CURRENT_DATE) as upcoming_sessions,
                (SELECT COUNT(*) FROM ticket WHERE purchase_date >= CURRENT_DATE) as today_tickets,
                (SELECT COALESCE(SUM(final_price), 0) FROM ticket WHERE purchase_date >= CURRENT_DATE) as today_revenue,
                (SELECT COUNT(*) FROM users WHERE status = 'Active') as active_users
        """
        result = query(sql)
        return result[0] if result else (0, 0, 0, 0, 0)